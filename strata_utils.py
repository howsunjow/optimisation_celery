from response_functions import Parameter, ResponseFunctionType
from openpyxl import load_workbook
import numpy as np

COSTS_ROW_START = 5
COSTS_COLUMN_START = 5
MODEL_VARIABLE_START_ROW = 8
MODEL_DATA_START_ROW = 19
CONSTRAINTS_ROW_START = 3
MODEL_DEFINITION_SHEET = 'Model Definition'

def get_functional_form_from_name(functionalform_name):
    """Given a functional form name found in the Excel export of a stratQED 
    model, this function returns the response function type

    Args:
        functional_form_name : str
            The functional form name

    Returns:
        ResponseFunctionType: The response function type corresponding to the functional form name           
    """
    functionalform_name_lower = functionalform_name.lower()
    if functionalform_name_lower.startswith('adbudg'):
        return ResponseFunctionType.ADBUDG
    elif functionalform_name_lower.startswith('linear'):
        return ResponseFunctionType.LINEAR
    elif functionalform_name_lower.startswith('logged'):
        return ResponseFunctionType.LOGGED
    elif functionalform_name_lower.startswith('exponential'):
        return ResponseFunctionType.EXPONENTIAL        
    elif functionalform_name_lower.startswith('arctan'):
        return ResponseFunctionType.ARCTAN
    elif functionalform_name_lower.startswith('power'):
        return ResponseFunctionType.POWER
    elif functionalform_name_lower.startswith('reciprocal'):
        return ResponseFunctionType.RECIPROCAL 


def get_column_with_variable(ws, avariable_name, arow=0):
    """ Finds the first column in the specified row containing the variable name. 

    Args:
        ws : object
            The excel worksheet
        avariable_name: str
            The name of the variable
        start_row: int
            The row to search

    Returns:
        str: The column containing the variable name          
    """    
    done = False
    found = False
    column_letter = None
    for col in ws.iter_cols(min_row=arow, max_row=arow, min_col=3):
        for cell in col:
            if cell.value == avariable_name:
                found = True
                column_letter = cell.column_letter
            elif not cell.value:
                done = True
                break
        if done or found:
            break
    return column_letter


def get_variable_parameters_from_sheet(ws, aVariableName, start_row=MODEL_VARIABLE_START_ROW):
    """Given a reference to an excel sheet, a variable name and a start row this function
    returns the parameters for the variable. This assumes that the variable parameters are
    stored in columns where the header is the variable name

    Args:
        ws: object
            The excel worksheet
        avariable_name: str
            The variable name
        start_row: int
            The row at which the variable block starts

    Returns:
        Parameter: The parameters of the variable i.e. response function type, carrover,
        coefficient, gamma, rho and lag           
    """    
    def get_cell_reference(acol, avalType):
        parameter_map = {'carryover': 9, 'response': 10, 'coefficient': 11,
                         'gamma': 12, 'rho': 13, 'lag': 14}
        if avalType in parameter_map:
            return acol + str(parameter_map[avalType])
    def get_float_or_default(avalue, default_value=0.0):
        try: 
            val = float(avalue)
            return val
        except ValueError:
            return default_value
    def get_int_or_default(avalue, default_value=0):
        try: 
            val = int(avalue)
            return val
        except ValueError:
            return default_value 
    column_letter = get_column_with_variable(ws, aVariableName, start_row)
    if column_letter:
        acell_ref = get_cell_reference(column_letter, 'response')
        aresponse = get_functional_form_from_name(ws[acell_ref].value)
        acell_ref = get_cell_reference(column_letter, 'coefficient')
        acoefficient = get_float_or_default(ws[acell_ref].value)
        acell_ref = get_cell_reference(column_letter, 'rho')
        arho = get_float_or_default(ws[acell_ref].value)
        acell_ref = get_cell_reference(column_letter, 'gamma')
        agamma = get_float_or_default(ws[acell_ref].value)
        acell_ref = get_cell_reference(column_letter, 'carryover')
        acarryover = get_float_or_default(ws[acell_ref].value)
        acell_ref = get_cell_reference(column_letter, 'lag')
        alag = get_int_or_default(ws[acell_ref].value)
        return Parameter(response=aresponse, coefficient=acoefficient, gamma=agamma,
                         rho=arho, carryover=acarryover, lag=alag)
    else:
        return None


def get_variable_parameters_from_excel_file(excel_filename, variable_names):
    wb = load_workbook(filename = excel_filename)
    ws = wb['Model Definition']
    return {avariablename: get_variable_parameters_from_sheet(ws, avariablename) for avariablename in variable_names}


def get_initial_adstock_values_from_excel_file(excel_filename, var_parameters):
    def get_initial_adstock_for_variable(ws, avariable_name, acarryover, start_row=MODEL_DATA_START_ROW):
        column_letter = get_column_with_variable(ws, avariable_name, start_row)
        # find the last row containing the data
        # for now we assume a single cross section so just use the max row
        if column_letter:
            ws_range = f'{column_letter}{start_row+1}:{column_letter}{ws.max_row}'
            #print(ws_range)
            avariable_units = [acost_cell[0].value for acost_cell in ws[ws_range]]
            return avariable_units[-1]
        return 0.0    
    wb = load_workbook(filename = excel_filename)
    ws = wb[MODEL_DEFINITION_SHEET]
    return {avariablename:
            get_initial_adstock_for_variable(ws, avariablename, var_parameters[avariablename].carryover)
            for avariablename in var_parameters}


def get_plan_costs_from_excel_file(plan_filename, start_row=COSTS_ROW_START, start_column=COSTS_COLUMN_START):
    wb = load_workbook(filename = plan_filename)
    ws = wb['Costs']
    # Find the last costs column
    done = False
    for col in ws.iter_cols(min_row=start_row, max_row=start_row, min_col=start_column):
        for cell in col:
            if not cell.value:
                done = True
                break
            else:
                column_letter = cell.column_letter
        if done:
            break
    # Return an empty dictionary if there is no recongnizable costs data on this sheet
    if not column_letter:
        return {};
    avariable_names = [acost_cell[0].value for acost_cell in ws[f'A{start_column}:A{ws.max_row}']]
    acosts_dict = {avar_name: 
                   [acost_cell.value for acost_cell in 
                    ws[f'E{avariable_names.index(avar_name)+start_row}:{column_letter}{avariable_names.index(avar_name)+start_row}'][0]]
                   for avar_name in avariable_names}
    return(acosts_dict)
         
def get_plan_constraints_from_excel_file(plan_filename, start_row=CONSTRAINTS_ROW_START):
    def get_float_or_default(avalue, default_value=0):
        try: 
            val = float(avalue)
            return val
        except ValueError:
            return default_value 
        except TypeError:
            return default_value    
    wb = load_workbook(filename = plan_filename)
    ws = wb['Constraints']
    avariable_names = [acost_cell[0].value for acost_cell in ws[f'A{start_row}:A{ws.max_row}']]
    aconstraint_dict = {avar: {'driver_miniumum': get_float_or_default(ws[f'B{start_row+i}'].value,0),
                                'driver_maximum': get_float_or_default(ws[f'C{start_row+i}'].value, np.inf),
                                'period_constraint_type': ws[f'E{start_row+i}'].value,
                                'period_maximum': get_float_or_default(ws[f'G{start_row+i}'].value, np.inf)} 
                                for i,avar in enumerate(avariable_names)}
    #print(aconstraint_dict)
    return aconstraint_dict
